#!/usr/bin/env python3
"""
Copie le mapping (MAPPING -> N) du BAL MODELE dans la colonne MAPPING du MODELE GL.xlsx.
Exécuter depuis la racine du projet : python backend/scripts/sync_gl_mapping.py
"""
import sys
from pathlib import Path

# Add backend to path (app is in backend/)
BACKEND = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BACKEND))

import openpyxl

from app.config.finance_excel_mapping import (
    BAL_MODELE_COLUMNS,
    BAL_MODELE_PATH,
    BAL_MODELE_SHEET,
    MODELE_GL_COLUMNS,
    MODELE_GL_PATH,
    MODELE_GL_SHEET,
)


def _normalize_account(val) -> str:
    """Normalise un compte (N ou Compte général) en chaîne pour comparaison."""
    if val is None:
        return ""
    if isinstance(val, (int, float)):
        return str(int(val))
    return str(val).strip()


def _build_n_to_mapping_from_bal() -> dict[str, str]:
    """Charge BAL MODELE et retourne {N: MAPPING} pour les comptes 8 chiffres."""
    n_to_mapping: dict[str, str] = {}
    if not BAL_MODELE_PATH.exists():
        return n_to_mapping
    wb = openpyxl.load_workbook(BAL_MODELE_PATH, read_only=True, data_only=True)
    try:
        if BAL_MODELE_SHEET not in wb.sheetnames:
            return n_to_mapping
        ws = wb[BAL_MODELE_SHEET]
        cols = BAL_MODELE_COLUMNS
        mapping_col = cols.get("mapping")
        if mapping_col is None:
            return n_to_mapping
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= mapping_col:
                continue
            n_val = row[cols["account"]]
            mapping_val = str(row[mapping_col] or "").strip()
            if not mapping_val:
                continue
            n_str = _normalize_account(n_val)
            if n_str:
                n_to_mapping[n_str] = mapping_val
    finally:
        wb.close()
    return n_to_mapping


def sync_gl_mapping() -> int:
    """
    Écrit le mapping du BAL dans la colonne MAPPING du MODELE GL.
    Retourne le nombre de cellules mises à jour.
    """
    n_to_mapping = _build_n_to_mapping_from_bal()
    if not n_to_mapping:
        print("Aucun mapping trouvé dans BAL MODELE.")
        return 0

    if not MODELE_GL_PATH.exists():
        print(f"Fichier introuvable : {MODELE_GL_PATH}")
        return 0

    wb = openpyxl.load_workbook(MODELE_GL_PATH, read_only=False)
    sheet_name = MODELE_GL_SHEET or wb.sheetnames[0]
    if sheet_name not in wb.sheetnames:
        print(f"Feuille '{sheet_name}' introuvable dans MODELE GL.")
        wb.close()
        return 0

    ws = wb[sheet_name]
    cols = MODELE_GL_COLUMNS
    account_col = cols["account"]
    mapping_col = cols["mapping"]

    updated = 0
    for row_idx in range(2, ws.max_row + 1):
        cell_account = ws.cell(row=row_idx, column=account_col + 1)
        compte = cell_account.value
        if compte is None:
            continue
        n_str = _normalize_account(compte)
        if n_str and n_str in n_to_mapping:
            cell_mapping = ws.cell(row=row_idx, column=mapping_col + 1)
            cell_mapping.value = n_to_mapping[n_str]
            updated += 1

    wb.save(MODELE_GL_PATH)
    wb.close()
    return updated


if __name__ == "__main__":
    count = sync_gl_mapping()
    print(f"MODELE GL.xlsx mis à jour : {count} écritures mappées.")
