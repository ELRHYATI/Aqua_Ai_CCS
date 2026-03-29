"""
Service Finance - Lecture des fichiers Excel et calcul des KPI YTD.

Lit MODELE RAPPORT.xlsx, BAL MODELE.xlsx et MODELE GL.xlsx.
- RAPPORT/BAL : données déjà agrégées (R, B, N-1).
- GL : écritures transactionnelles, agrégées par compte pour YTD Réalisé, fusion avec BAL pour B et N-1.
"""

import re
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Optional

import openpyxl

from app.config.finance_excel_mapping import (
    BAL_MODELE_COLUMNS,
    BAL_MODELE_PATH,
    BAL_MODELE_SHEET,
    MODELE_GL_COLUMNS,
    MODELE_GL_PATH,
    MODELE_GL_SHEET,
    MODELE_RAPPORT_COLUMNS,
    MODELE_RAPPORT_PATH,
    MODELE_RAPPORT_PERIOD_HEADER_INDEX,
    MODELE_RAPPORT_SHEET,
    RAPPORT_TO_SAP,
)


@dataclass
class FinanceRow:
    """Ligne brute issue du mapping Excel."""

    account: str
    label: str
    year: int
    month: int
    actual: float
    budget: float
    last_year: float


@dataclass
class FinanceKpiRow:
    """KPI calculé pour une ligne/compte."""

    account: str
    label: str
    ytd: float
    budget_ytd: float
    last_year_ytd: float
    var_budget: Optional[float]  # (R - B) / B, None si division par zéro
    var_last_year: Optional[float]  # (R - N-1) / N-1, None si division par zéro
    var_budget_div_zero: bool
    var_last_year_div_zero: bool


def _safe_float(val) -> float:
    """Convertit une valeur en float, 0 si invalide."""
    if val is None or val == "":
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", ".")
    if not s or s.upper() in ("#VALUE!", "#REF!", "#DIV/0!", "NC", "N/A"):
        return 0.0
    try:
        return float(s)
    except (ValueError, TypeError, InvalidOperation):
        return 0.0


def _safe_str(val) -> str:
    """Convertit en chaîne non vide."""
    if val is None:
        return ""
    return str(val).strip()


def compute_var_budget(actual: float, budget: float) -> tuple[Optional[float], bool]:
    """
    Calcule (R - B) / B.
    Returns: (valeur ou None si B=0, flag division_par_zero)
    """
    if budget == 0:
        return None, True
    return (actual - budget) / budget, False


def compute_var_last_year(actual: float, last_year: float) -> tuple[Optional[float], bool]:
    """
    Calcule (R - N-1) / N-1.
    Returns: (valeur ou None si N-1=0, flag division_par_zero)
    """
    if last_year == 0:
        return None, True
    return (actual - last_year) / last_year, False


def _parse_period_from_header(header_cell: str) -> tuple[int, int]:
    """
    Parse l'année et le mois depuis un en-tête type "PERIODE YTD 01 2026".
    Retourne (year, month). Par défaut 2026, 1.
    """
    year, month = 2026, 1
    if not header_cell:
        return year, month
    s = str(header_cell).upper()
    # Chercher pattern MM AAAA ou MM-AAAA
    m = re.search(r"(\d{1,2})\s*[/\-]?\s*(\d{4})", s)
    if m:
        month = int(m.group(1))
        year = int(m.group(2))
    return year, month


def _load_modele_rapport(path: Path, year_filter: Optional[int] = None) -> list[FinanceRow]:
    """Charge MODELE RAPPORT.xlsx et retourne les lignes mappées."""
    rows_out: list[FinanceRow] = []
    if not path.exists():
        return rows_out

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if MODELE_RAPPORT_SHEET not in wb.sheetnames:
            return rows_out
        ws = wb[MODELE_RAPPORT_SHEET]
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        year, month = 2026, 1
        if header_row and len(header_row[0]) > MODELE_RAPPORT_PERIOD_HEADER_INDEX:
            year, month = _parse_period_from_header(
                str(header_row[0][MODELE_RAPPORT_PERIOD_HEADER_INDEX] or "")
            )

        if year_filter is not None and year != year_filter:
            return rows_out

        cols = MODELE_RAPPORT_COLUMNS
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[cols["account"]] is None:
                # Ligne de sous-total (ex: P1119T) : on garde si label présent
                if len(row) > cols["label"] and row[cols["label"]]:
                    account = _safe_str(row[cols["account"]]) or _safe_str(row[cols["label"]])[:20]
                else:
                    continue
            else:
                account = _safe_str(row[cols["account"]])

            label = _safe_str(row[cols["label"]]) if len(row) > cols["label"] else ""
            actual = _safe_float(row[cols["actual"]]) if len(row) > cols["actual"] else 0.0
            budget = _safe_float(row[cols["budget"]]) if len(row) > cols["budget"] else 0.0
            last_year = _safe_float(row[cols["last_year"]]) if len(row) > cols["last_year"] else 0.0

            # Ignorer lignes toutes à zéro sans label utile
            if not account and not label:
                continue

            rows_out.append(
                FinanceRow(
                    account=account or f"L{len(rows_out)+1}",
                    label=label or account,
                    year=year,
                    month=month,
                    actual=actual,
                    budget=budget,
                    last_year=last_year,
                )
            )
    finally:
        wb.close()

    return rows_out


def _parse_gl_date(val) -> Optional[tuple[int, int]]:
    """
    Parse la date comptable GL (format DD.MM.YYYY ou date Excel).
    Retourne (year, month) ou None si invalide.
    """
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.year, val.month
    if isinstance(val, (int, float)):
        try:
            from datetime import timedelta
            d = datetime(1899, 12, 30) + timedelta(days=float(val))
            return d.year, d.month
        except (TypeError, ValueError):
            return None
    s = str(val).strip()
    m = re.match(r"(\d{1,2})\.(\d{1,2})\.(\d{4})", s)
    if m:
        return int(m.group(3)), int(m.group(2))
    return None


def _load_modele_gl(
    path: Path,
    year_filter: Optional[int] = None,
    month_to: Optional[int] = None,
) -> dict[str, tuple[float, str]]:
    """
    Charge MODELE GL.xlsx (écritures transactionnelles) et agrège par compte.
    Retourne {account: (ytd_actual, label)}.
    year_filter: filtrer par année. month_to: inclure les mois 1..month_to (None = tous).
    """
    agg: dict[str, tuple[float, str]] = {}
    if not path.exists():
        return agg

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_name = MODELE_GL_SHEET or wb.sheetnames[0]
        if sheet_name not in wb.sheetnames:
            return agg
        ws = wb[sheet_name]
        cols = MODELE_GL_COLUMNS
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= cols["amount"]:
                continue
            account_val = row[cols["account"]]
            if account_val is None:
                continue
            account = _safe_str(account_val)
            if not account:
                continue
            label = _safe_str(row[cols["label"]]) if len(row) > cols["label"] else account
            amount = _safe_float(row[cols["amount"]])
            date_val = row[cols["date"]] if len(row) > cols["date"] else None
            parsed = _parse_gl_date(date_val)
            if parsed:
                year, month = parsed
                if year_filter is not None and year != year_filter:
                    continue
                if month_to is not None and month > month_to:
                    continue
            # Pas de date valide = on inclut quand même (année par défaut)
            elif year_filter is not None:
                continue

            prev_sum, prev_label = agg.get(account, (0.0, label))
            agg[account] = (prev_sum + amount, prev_label or label)
    finally:
        wb.close()

    return agg


def _resolve_account_to_sap_list(account: str) -> list[str]:
    """
    Résout un code court (P1131, E1110) vers la liste des comptes SAP 8 chiffres.
    Utilise RAPPORT_TO_SAP (config), BAL MODELE (MAPPING/BILAN_CPC -> N), et matching par libellé.
    """
    account_str = str(account).strip()
    if not account_str:
        return []

    # Déjà 8 chiffres (SAP) : utiliser tel quel
    if account_str.isdigit() and len(account_str) == 8:
        return [account_str]

    # 1. Override config (RAPPORT_TO_SAP)
    if account_str in RAPPORT_TO_SAP:
        sap_list = [str(s).strip() for s in RAPPORT_TO_SAP[account_str] if s]
        if sap_list:
            return sap_list

    # 2. BAL MODELE : MAPPING ou BILAN_CPC -> N (issu des fichiers déposés)
    if not BAL_MODELE_PATH.exists():
        return []
    try:
        sap_accounts: list[str] = []
        wb = openpyxl.load_workbook(BAL_MODELE_PATH, read_only=True, data_only=True)
        try:
            if BAL_MODELE_SHEET not in wb.sheetnames:
                return []
            ws = wb[BAL_MODELE_SHEET]
            cols = BAL_MODELE_COLUMNS
            mapping_col = cols.get("mapping")
            bilan_cpc_col = cols.get("bilan_cpc")
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) <= cols["account"]:
                    continue
                # MAPPING (ex: E1110)
                if mapping_col is not None and len(row) > mapping_col:
                    mapping_val = str(row[mapping_col] or "").strip()
                    if mapping_val.upper() == account_str.upper():
                        n_val = row[cols["account"]]
                        if n_val is not None:
                            sap = str(int(n_val)) if isinstance(n_val, (int, float)) else str(n_val).strip()
                            if sap.isdigit() and len(sap) == 8:
                                sap_accounts.append(sap)
                                break
                # BILAN_CPC (ex: P1110) si présent dans le BAL
                if not sap_accounts and bilan_cpc_col is not None and len(row) > bilan_cpc_col:
                    bilan_val = str(row[bilan_cpc_col] or "").strip()
                    if bilan_val.upper() == account_str.upper():
                        n_val = row[cols["account"]]
                        if n_val is not None:
                            sap = str(int(n_val)) if isinstance(n_val, (int, float)) else str(n_val).strip()
                            if sap.isdigit() and len(sap) == 8:
                                sap_accounts.append(sap)
                                break
        finally:
            wb.close()
        if sap_accounts:
            return sap_accounts

        # 3. Matching par libellé : RAPPORT (P1110) -> BAL par similitude du libellé
        rapport_label = _get_rapport_label_for_account(account_str)
        if rapport_label:
            sap_from_label = _resolve_sap_from_bal_by_label(rapport_label)
            if sap_from_label:
                return sap_from_label
    except Exception:
        pass
    return []


def _get_rapport_label_for_account(account: str) -> Optional[str]:
    """Retourne le libellé du compte depuis MODELE RAPPORT si trouvé."""
    if not MODELE_RAPPORT_PATH.exists():
        return None
    try:
        wb = openpyxl.load_workbook(MODELE_RAPPORT_PATH, read_only=True, data_only=True)
        try:
            if MODELE_RAPPORT_SHEET not in wb.sheetnames:
                return None
            ws = wb[MODELE_RAPPORT_SHEET]
            cols = MODELE_RAPPORT_COLUMNS
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) <= cols["account"]:
                    continue
                acc_val = _safe_str(row[cols["account"]])
                if acc_val.upper() == account.upper():
                    return _safe_str(row[cols["label"]]) if len(row) > cols["label"] else None
        finally:
            wb.close()
    except Exception:
        pass
    return None


def _resolve_sap_from_bal_by_label(label: str) -> list[str]:
    """Cherche dans BAL les comptes dont COMPTE contient le libellé; retourne leurs N (8 chiffres)."""
    if not label or not BAL_MODELE_PATH.exists():
        return []
    label_lower = label.lower().strip()
    if len(label_lower) < 3:
        return []
    results: list[str] = []
    try:
        wb = openpyxl.load_workbook(BAL_MODELE_PATH, read_only=True, data_only=True)
        try:
            if BAL_MODELE_SHEET not in wb.sheetnames:
                return []
            ws = wb[BAL_MODELE_SHEET]
            cols = BAL_MODELE_COLUMNS
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) <= cols["label"]:
                    continue
                compte_val = _safe_str(row[cols["label"]])
                if label_lower in compte_val.lower():
                    n_val = row[cols["account"]]
                    if n_val is not None:
                        sap = str(int(n_val)) if isinstance(n_val, (int, float)) else str(n_val).strip()
                        if sap.isdigit() and len(sap) == 8:
                            results.append(sap)
        finally:
            wb.close()
    except Exception:
        pass
    return results


def get_gl_entries_for_account(
    account: str,
    year: Optional[int] = None,
) -> list[dict]:
    """
    Return raw GL line entries for a given account (for drill-down).
    Each entry: {date_str, label, amount, year, month}.
    """
    entries: list[dict] = []
    if not account or not MODELE_GL_PATH.exists():
        return entries

    wb = openpyxl.load_workbook(MODELE_GL_PATH, read_only=True, data_only=True)
    try:
        sheet_name = MODELE_GL_SHEET or wb.sheetnames[0]
        if sheet_name not in wb.sheetnames:
            return entries
        ws = wb[sheet_name]
        cols = MODELE_GL_COLUMNS
        account_str = str(account).strip()

        # Résoudre code court (P1131, E1110) -> liste comptes SAP 8 chiffres
        sap_accounts = _resolve_account_to_sap_list(account_str)
        if not sap_accounts:
            sap_accounts = [account_str]  # fallback: match exact

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= cols["amount"]:
                continue
            row_account = str(row[cols["account"]] or "").strip()
            row_mapping = str(row[cols.get("mapping", cols["account"])] or "").strip() if "mapping" in cols else ""
            if row_account not in sap_accounts and row_mapping != account_str:
                continue
            label = _safe_str(row[cols["label"]]) if len(row) > cols["label"] else ""
            amount = _safe_float(row[cols["amount"]])
            date_val = row[cols["date"]] if len(row) > cols["date"] else None
            date_str = str(date_val) if date_val else ""
            parsed = _parse_gl_date(date_val)
            if parsed and year is not None and parsed[0] != year:
                continue
            entries.append({
                "date_str": date_str,
                "label": label,
                "amount": amount,
                "year": parsed[0] if parsed else None,
                "month": parsed[1] if parsed else None,
            })
    finally:
        wb.close()

    return sorted(entries, key=lambda x: (x.get("year") or 0, x.get("month") or 0, x.get("date_str", "")))


def get_rapport_row_for_account(
    account: str,
    year: Optional[int] = None,
) -> Optional[FinanceKpiRow]:
    """
    Retourne les KPI du compte depuis MODELE RAPPORT (fichier déposé).
    Utile pour générer un commentaire IA basé sur les données RAPPORT quand le GL n'a pas d'écritures.
    """
    if not account or not MODELE_RAPPORT_PATH.exists():
        return None
    account_str = str(account).strip().upper()
    rapport_rows = _load_modele_rapport(MODELE_RAPPORT_PATH, year_filter=year)
    for r in rapport_rows:
        if _safe_str(r.account).upper() == account_str:
            kpis = compute_kpis([r])
            return kpis[0] if kpis else None
    return None


def _load_bal_modele(path: Path, year_filter: Optional[int] = None) -> list[FinanceRow]:
    """
    Charge BAL MODELE.xlsx. Les colonnes R YTD, B YTD, LY YTD contiennent déjà les YTD.
    year_filter : on suppose year=2026 pour BAL si non précisé dans le fichier.
    """
    rows_out: list[FinanceRow] = []
    if not path.exists():
        return rows_out

    # Par défaut BAL n'a pas year/month dans le fichier, on utilise l'année courante
    year = year_filter if year_filter is not None else 2026
    month = 12  # YTD complet

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if BAL_MODELE_SHEET not in wb.sheetnames:
            return rows_out
        ws = wb[BAL_MODELE_SHEET]
        cols = BAL_MODELE_COLUMNS
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[cols["account"]] is None:
                continue
            account = _safe_str(row[cols["account"]])
            if not account:
                continue
            label = _safe_str(row[cols["label"]]) if len(row) > cols["label"] else account
            actual = _safe_float(row[cols["actual"]]) if len(row) > cols["actual"] else 0.0
            budget = _safe_float(row[cols["budget"]]) if len(row) > cols["budget"] else 0.0
            last_year = _safe_float(row[cols["last_year"]]) if len(row) > cols["last_year"] else 0.0

            rows_out.append(
                FinanceRow(
                    account=account,
                    label=label,
                    year=year,
                    month=month,
                    actual=actual,
                    budget=budget,
                    last_year=last_year,
                )
            )
    finally:
        wb.close()

    return rows_out


def sync_gl_mapping() -> int:
    """
    Copie le mapping (MAPPING -> N) du BAL MODELE dans la colonne MAPPING du MODELE GL.
    Retourne le nombre de cellules mises à jour.
    """
    n_to_mapping: dict[str, str] = {}
    if not BAL_MODELE_PATH.exists():
        return 0
    mapping_col = BAL_MODELE_COLUMNS.get("mapping")
    if mapping_col is None:
        return 0

    wb = openpyxl.load_workbook(BAL_MODELE_PATH, read_only=True, data_only=True)
    try:
        if BAL_MODELE_SHEET not in wb.sheetnames:
            return 0
        ws = wb[BAL_MODELE_SHEET]
        cols = BAL_MODELE_COLUMNS
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= mapping_col:
                continue
            n_val = row[cols["account"]]
            mapping_val = str(row[mapping_col] or "").strip()
            if not mapping_val:
                continue
            n_str = str(int(n_val)) if isinstance(n_val, (int, float)) else str(n_val or "").strip()
            if n_str:
                n_to_mapping[n_str] = mapping_val
    finally:
        wb.close()

    if not n_to_mapping or not MODELE_GL_PATH.exists():
        return 0

    wb = openpyxl.load_workbook(MODELE_GL_PATH, read_only=False)
    sheet_name = MODELE_GL_SHEET or wb.sheetnames[0]
    if sheet_name not in wb.sheetnames:
        wb.close()
        return 0
    ws = wb[sheet_name]
    cols = MODELE_GL_COLUMNS
    updated = 0
    for row_idx in range(2, ws.max_row + 1):
        cell_account = ws.cell(row=row_idx, column=cols["account"] + 1)
        compte = cell_account.value
        if compte is None:
            continue
        n_str = str(int(compte)) if isinstance(compte, (int, float)) else str(compte or "").strip()
        if n_str and n_str in n_to_mapping:
            ws.cell(row=row_idx, column=cols["mapping"] + 1).value = n_to_mapping[n_str]
            updated += 1
    wb.save(MODELE_GL_PATH)
    wb.close()
    return updated


def load_finance_rows(
    year: Optional[int] = None,
    *,
    prefer_rapport: bool = True,
    use_gl: bool = False,
    month_to: Optional[int] = None,
) -> list[FinanceRow]:
    """
    Charge les lignes finance depuis les fichiers Excel.
    - prefer_rapport=True : MODELE RAPPORT en priorité (CPC avec R, B, N-1).
    - prefer_rapport=False : BAL MODELE (R YTD, B YTD, LY YTD).
    - use_gl=True : MODELE GL pour R (agrégé par compte), fusion avec BAL pour B et N-1.
      Utile quand on veut le Réalisé depuis le Grand Livre.
    - month_to : pour GL, inclure uniquement les mois 1..month_to (None = tous).
    """
    if use_gl:
        gl_agg = _load_modele_gl(MODELE_GL_PATH, year_filter=year, month_to=month_to)
        bal_rows = _load_bal_modele(BAL_MODELE_PATH, year_filter=year)
        # Merge : R depuis GL, B et N-1 depuis BAL (même numéro de compte)
        bal_by_account: dict[str, FinanceRow] = {r.account: r for r in bal_rows}
        year_val = year or 2026
        month_val = month_to or 12
        rows_out: list[FinanceRow] = []
        for account, (ytd_actual, label) in gl_agg.items():
            bal = bal_by_account.get(account)
            budget = bal.budget if bal else 0.0
            last_year = bal.last_year if bal else 0.0
            if bal and bal.label:
                label = bal.label
            rows_out.append(
                FinanceRow(
                    account=account,
                    label=label or account,
                    year=year_val,
                    month=month_val,
                    actual=ytd_actual,
                    budget=budget,
                    last_year=last_year,
                )
            )
        return rows_out

    rapport_rows = _load_modele_rapport(MODELE_RAPPORT_PATH, year_filter=year)
    bal_rows = _load_bal_modele(BAL_MODELE_PATH, year_filter=year)

    if prefer_rapport and rapport_rows:
        return rapport_rows
    if bal_rows:
        return bal_rows
    return rapport_rows


def compute_kpis(rows: list[FinanceRow]) -> list[FinanceKpiRow]:
    """
    Pour chaque ligne, les valeurs R, B, N-1 sont déjà YTD (depuis le fichier).
    Calcule varBudget et varLastYear avec gestion division par zéro.
    """
    result: list[FinanceKpiRow] = []
    for r in rows:
        var_budget, div_b = compute_var_budget(r.actual, r.budget)
        var_ly, div_ly = compute_var_last_year(r.actual, r.last_year)
        result.append(
            FinanceKpiRow(
                account=r.account,
                label=r.label,
                ytd=r.actual,
                budget_ytd=r.budget,
                last_year_ytd=r.last_year,
                var_budget=var_budget,
                var_last_year=var_ly,
                var_budget_div_zero=div_b,
                var_last_year_div_zero=div_ly,
            )
        )
    return result


def get_finance_kpis(
    year: Optional[int] = None,
    *,
    source: str = "rapport",
    month_to: Optional[int] = None,
) -> list[FinanceKpiRow]:
    """
    Point d'entrée principal : charge les données et retourne les KPI.
    source: "rapport" (défaut), "bal" ou "gl".
    - rapport : MODELE RAPPORT (CPC, R/B/N-1 déjà agrégés)
    - bal : BAL MODELE (R YTD, B YTD, LY YTD)
    - gl : MODELE GL (R agrégé depuis écritures) + BAL pour B et N-1
    """
    use_gl = source == "gl"
    prefer_rapport = source == "rapport"
    rows = load_finance_rows(
        year=year,
        prefer_rapport=prefer_rapport,
        use_gl=use_gl,
        month_to=month_to,
    )
    return compute_kpis(rows)


def get_finance_kpis_summary(
    year: Optional[int] = None,
    *,
    source: str = "rapport",
    month_to: Optional[int] = None,
) -> dict:
    """
    Retourne les KPI agrégés (totaux) + liste des lignes.
    """
    kpis = get_finance_kpis(year=year, source=source, month_to=month_to)
    total_ytd = sum(k.ytd for k in kpis)
    total_budget = sum(k.budget_ytd for k in kpis)
    total_last_year = sum(k.last_year_ytd for k in kpis)
    var_budget, _ = compute_var_budget(total_ytd, total_budget)
    var_last_year, _ = compute_var_last_year(total_ytd, total_last_year)
    return {
        "total_ytd": total_ytd,
        "total_budget_ytd": total_budget,
        "total_last_year_ytd": total_last_year,
        "var_budget_pct": var_budget,
        "var_last_year_pct": var_last_year,
        "rows": kpis,
    }
