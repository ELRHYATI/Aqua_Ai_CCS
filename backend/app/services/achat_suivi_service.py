"""
Achat Suivi service: parse 'Suivi Global CCS.xlsm' and compute KPIs.
Reads from data/ folder, computes DA/BC counts, statuses, timelines, and KPIs
matching the Excel KPIs sheet structure.
"""

from datetime import date, datetime
from pathlib import Path
from typing import Any
from collections import Counter, defaultdict

import openpyxl

# Column indices for "Suivi Global" sheet (0-based, row 2 = headers, row 3+ = data)
COL_TYPE_PROCESS = 0   # Achat / Appro
COL_CAPEX_OPEX = 1     # Capex / Opex
COL_ID_DA = 2
COL_DATE_CREATION_DA = 3
COL_TYPE_DA = 4
COL_ID_PRODUIT = 5
COL_PRODUIT = 6
COL_CATEGORIE = 7
COL_QTE_DA = 8
COL_DEMANDEUR = 9
COL_STATUT_DA = 10
COL_CODE_CAPEX = 11
COL_ID_CDE = 12
COL_POSTE = 13
COL_DATE_CREATION_BC = 14
COL_ACHETEUR = 15
COL_STATUT_CDE = 16
COL_APPROBATEUR = 17
COL_QTE_BC = 18
COL_PRIX = 19
COL_VALEUR = 20
COL_FOURNISSEUR = 21
COL_DATE_DEBUT_LIVRAISON = 22
COL_BL = 23
COL_ID_RECEPTION = 24
COL_MAGASINIER = 25
COL_DATE_RECEPTION = 26
COL_QTE_RECEPTION = 27
COL_FACTURE = 28
COL_COMMENTAIRE = 29


def _find_excel() -> Path | None:
    candidates = [
        Path(__file__).resolve().parent.parent.parent.parent / "data" / "Suivi Global CCS.xlsm",
        Path.cwd().parent / "data" / "Suivi Global CCS.xlsm",
        Path.cwd() / "data" / "Suivi Global CCS.xlsm",
    ]
    for p in candidates:
        if p.exists():
            return p
    return None


def _safe_float(val) -> float | None:
    if val is None or val == "":
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _safe_date(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        d = val.date() if isinstance(val, datetime) else val
        return d.isoformat()
    return None


def _safe_str(val) -> str | None:
    if val is None or val == "":
        return None
    return str(val).strip()


def _parse_iso_date(s: str | None) -> date | None:
    if not s:
        return None
    try:
        return date.fromisoformat(s[:10])
    except (ValueError, TypeError):
        return None


def compute_lead_time_and_service_kpis(data: list[dict]) -> dict[str, Any]:
    """
    - Délai moyen de traitement DA : jours entre date_creation_da et date_creation_bc (même ligne).
    - Taux OTD (On Time Delivery) : parmi les BC avec date de réception et date début livraison promise,
      part reçue au plus tard à la date promise.
    - Taux OTIF (On Time In Full) : parmi les lignes éligibles (promesse + réception + qtés),
      part à la fois à l'heure et avec qte_reception >= qte_bc.
    """
    if not data:
        return {
            "delai_moyen_traitement_da_jours": None,
            "delai_traitement_da_echantillon_n": 0,
            "taux_otd_pct": None,
            "bc_otd_echantillon_n": 0,
            "taux_otif_pct": None,
            "bc_otif_echantillon_n": 0,
        }

    lead_times: list[int] = []
    for r in data:
        d_da = _parse_iso_date(r.get("date_creation_da"))
        d_bc = _parse_iso_date(r.get("date_creation_bc"))
        if d_da and d_bc:
            days = (d_bc - d_da).days
            if days >= 0:
                lead_times.append(days)

    avg_lead = round(sum(lead_times) / len(lead_times), 1) if lead_times else None

    otd_ok = 0
    otd_tot = 0
    otif_ok = 0
    otif_tot = 0

    for r in data:
        rec = _parse_iso_date(r.get("date_reception"))
        if not rec:
            continue
        prom = _parse_iso_date(r.get("date_debut_livraison"))
        if prom:
            otd_tot += 1
            if rec <= prom:
                otd_ok += 1

        qb = r.get("qte_bc")
        qr = r.get("qte_reception")
        if prom and qb is not None and qr is not None:
            try:
                fqb = float(qb)
                fqr = float(qr)
            except (TypeError, ValueError):
                continue
            if fqb <= 0:
                continue
            otif_tot += 1
            on_time = rec <= prom
            full = fqr >= fqb * 0.999
            if on_time and full:
                otif_ok += 1

    otd_pct = round(100.0 * otd_ok / otd_tot, 1) if otd_tot else None
    otif_pct = round(100.0 * otif_ok / otif_tot, 1) if otif_tot else None

    return {
        "delai_moyen_traitement_da_jours": avg_lead,
        "delai_traitement_da_echantillon_n": len(lead_times),
        "taux_otd_pct": otd_pct,
        "bc_otd_echantillon_n": otd_tot,
        "taux_otif_pct": otif_pct,
        "bc_otif_echantillon_n": otif_tot,
    }


def load_suivi_data(excel_path: Path | None = None) -> list[dict]:
    path = excel_path or _find_excel()
    if not path or not path.exists():
        return []

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if "Suivi Global" not in wb.sheetnames:
            return []
        ws = wb["Suivi Global"]
        rows = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[COL_TYPE_PROCESS] is None:
                continue
            rows.append({
                "type_process": _safe_str(row[COL_TYPE_PROCESS]),
                "capex_opex": _safe_str(row[COL_CAPEX_OPEX]),
                "id_da": row[COL_ID_DA],
                "date_creation_da": _safe_date(row[COL_DATE_CREATION_DA]),
                "type_da": _safe_str(row[COL_TYPE_DA]),
                "id_produit": row[COL_ID_PRODUIT],
                "produit": _safe_str(row[COL_PRODUIT]),
                "categorie": _safe_str(row[COL_CATEGORIE]),
                "qte_da": _safe_float(row[COL_QTE_DA]),
                "demandeur": _safe_str(row[COL_DEMANDEUR]),
                "statut_da": _safe_str(row[COL_STATUT_DA]),
                "code_capex": _safe_str(row[COL_CODE_CAPEX]),
                "id_cde": row[COL_ID_CDE],
                "poste": row[COL_POSTE],
                "date_creation_bc": _safe_date(row[COL_DATE_CREATION_BC]),
                "acheteur": _safe_str(row[COL_ACHETEUR]),
                "statut_cde": _safe_str(row[COL_STATUT_CDE]),
                "approbateur": _safe_str(row[COL_APPROBATEUR]),
                "qte_bc": _safe_float(row[COL_QTE_BC]),
                "prix": _safe_float(row[COL_PRIX]),
                "valeur": _safe_float(row[COL_VALEUR]),
                "fournisseur": _safe_str(row[COL_FOURNISSEUR]),
                "date_debut_livraison": _safe_date(row[COL_DATE_DEBUT_LIVRAISON]),
                "bl": _safe_str(row[COL_BL]) if len(row) > COL_BL else None,
                "id_reception": _safe_str(row[COL_ID_RECEPTION]) if len(row) > COL_ID_RECEPTION else None,
                "magasinier": _safe_str(row[COL_MAGASINIER]) if len(row) > COL_MAGASINIER else None,
                "date_reception": _safe_date(row[COL_DATE_RECEPTION]) if len(row) > COL_DATE_RECEPTION else None,
                "qte_reception": _safe_float(row[COL_QTE_RECEPTION]) if len(row) > COL_QTE_RECEPTION else None,
                "facture": _safe_str(row[COL_FACTURE]) if len(row) > COL_FACTURE else None,
                "commentaire": _safe_str(row[COL_COMMENTAIRE]) if len(row) > COL_COMMENTAIRE else None,
            })
        return rows
    finally:
        wb.close()


def compute_summary(data: list[dict]) -> dict[str, Any]:
    """Compute high-level summary: DA en cours, BC en cours, etc."""
    if not data:
        return {"total_lignes": 0, "da_en_cours": 0, "bc_en_cours": 0, "bc_livrees": 0,
                "valeur_totale": 0, "fournisseurs": 0, "categories": 0}

    unique_da = set()
    da_en_cours = set()
    bc_en_cours = set()
    bc_livrees = set()
    valeur_totale = 0.0
    fournisseurs = set()
    categories = set()

    for r in data:
        da_id = r["id_da"]
        cde_id = r["id_cde"]
        statut_da = (r["statut_da"] or "").lower()
        statut_cde = (r["statut_cde"] or "").lower()

        if da_id is not None:
            unique_da.add(da_id)
        if r["fournisseur"]:
            fournisseurs.add(r["fournisseur"])
        if r["categorie"]:
            categories.add(r["categorie"])
        if r["valeur"]:
            valeur_totale += r["valeur"]

        # DA "en cours" = statut DA is "Aucun document lié" (no BC yet)
        if "aucun" in statut_da and da_id is not None:
            da_en_cours.add(da_id)

        # BC statuses
        if cde_id is not None:
            if "approbation" in statut_cde or "préparation" in statut_cde or "pr\xe9paration" in statut_cde or "envoy" in statut_cde or "confirmation" in statut_cde or "r\xe9vision" in statut_cde:
                bc_en_cours.add(cde_id)
            elif "li\xe9" in statut_cde or "document" in statut_cde:
                # "Document lié créé" = delivered/processed
                if r["bl"] == "Oui" or r["date_reception"]:
                    bc_livrees.add(cde_id)
                else:
                    bc_en_cours.add(cde_id)

    return {
        "total_lignes": len(data),
        "total_da": len(unique_da),
        "da_en_cours": len(da_en_cours),
        "bc_en_cours": len(bc_en_cours),
        "bc_livrees": len(bc_livrees),
        "valeur_totale": round(valeur_totale, 2),
        "fournisseurs": len(fournisseurs),
        "categories": len(categories),
    }


def compute_kpis(data: list[dict]) -> dict[str, Any]:
    """
    Compute KPIs matching the Excel KPIs sheet:
    DA: en attente approbation, consultation en cours, DA traité, autres
    Commandes: en cours approbation, confirmée, envoyée au fournisseur, livrée, autres
    Magasin: livraison sans BC, partielle, non conforme, retour fournisseur, sans BL, livrée, autres
    Grouped by Opex/Capex and by Responsable.
    """
    # Count by statut_da
    statut_da_counts = Counter()
    statut_cde_counts = Counter()
    by_capex_opex = defaultdict(lambda: {"da": Counter(), "cde": Counter(), "valeur": 0.0, "count": 0})
    by_demandeur = defaultdict(lambda: {"da": Counter(), "cde": Counter(), "count": 0})
    by_categorie = defaultdict(lambda: {"count": 0, "valeur": 0.0})
    by_fournisseur = defaultdict(lambda: {"count": 0, "valeur": 0.0})
    by_month = defaultdict(lambda: {"da_created": 0, "bc_created": 0, "valeur": 0.0})

    for r in data:
        s_da = r["statut_da"] or "Inconnu"
        s_cde = r["statut_cde"] or "Sans BC"
        co = r["capex_opex"] or "Inconnu"
        dem = r["demandeur"] or "Inconnu"
        cat = r["categorie"] or "Autre"
        frn = r["fournisseur"] or "Inconnu"
        val = r["valeur"] or 0

        statut_da_counts[s_da] += 1
        statut_cde_counts[s_cde] += 1

        by_capex_opex[co]["da"][s_da] += 1
        by_capex_opex[co]["cde"][s_cde] += 1
        by_capex_opex[co]["valeur"] += val
        by_capex_opex[co]["count"] += 1

        by_demandeur[dem]["da"][s_da] += 1
        by_demandeur[dem]["cde"][s_cde] += 1
        by_demandeur[dem]["count"] += 1

        by_categorie[cat]["count"] += 1
        by_categorie[cat]["valeur"] += val

        by_fournisseur[frn]["count"] += 1
        by_fournisseur[frn]["valeur"] += val

        # Timeline by month
        d = r["date_creation_da"]
        if d:
            month_key = d[:7]  # "YYYY-MM"
            by_month[month_key]["da_created"] += 1
        d2 = r["date_creation_bc"]
        if d2:
            month_key2 = d2[:7]
            by_month[month_key2]["bc_created"] += 1
            by_month[month_key2]["valeur"] += val

    # Top categories and fournisseurs
    top_categories = sorted(by_categorie.items(), key=lambda x: x[1]["valeur"], reverse=True)[:15]
    top_fournisseurs = sorted(by_fournisseur.items(), key=lambda x: x[1]["valeur"], reverse=True)[:15]

    # Timeline sorted
    timeline = [{"month": k, **v} for k, v in sorted(by_month.items())]

    # By capex/opex
    capex_opex_summary = {}
    for co, d in by_capex_opex.items():
        capex_opex_summary[co] = {
            "count": d["count"],
            "valeur": round(d["valeur"], 2),
            "statut_da": dict(d["da"]),
            "statut_cde": dict(d["cde"]),
        }

    # By demandeur (top 10)
    top_demandeurs = sorted(by_demandeur.items(), key=lambda x: x[1]["count"], reverse=True)[:10]
    demandeur_summary = [
        {"name": name, "count": d["count"], "statut_da": dict(d["da"]), "statut_cde": dict(d["cde"])}
        for name, d in top_demandeurs
    ]

    return {
        "statut_da": dict(statut_da_counts),
        "statut_cde": dict(statut_cde_counts),
        "capex_opex": capex_opex_summary,
        "by_demandeur": demandeur_summary,
        "top_categories": [{"name": n, "count": d["count"], "valeur": round(d["valeur"], 2)} for n, d in top_categories],
        "top_fournisseurs": [{"name": n, "count": d["count"], "valeur": round(d["valeur"], 2)} for n, d in top_fournisseurs],
        "timeline": timeline,
    }


def get_achat_suivi_full(excel_path: Path | None = None) -> dict[str, Any]:
    """Main entry: returns summary, KPIs, and recent records."""
    data = load_suivi_data(excel_path)
    if not data:
        empty_summary = compute_summary([])
        empty_summary.update(compute_lead_time_and_service_kpis([]))
        return {"summary": empty_summary, "kpis": compute_kpis([]), "records": []}

    summary = compute_summary(data)
    summary.update(compute_lead_time_and_service_kpis(data))
    kpis = compute_kpis(data)

    # Return last 100 records for the table (most recent first)
    records_sorted = sorted(data, key=lambda r: r.get("date_creation_da") or "", reverse=True)
    return {
        "summary": summary,
        "kpis": kpis,
        "records": records_sorted[:200],
    }
