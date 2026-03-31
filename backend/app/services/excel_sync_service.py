"""
Excel sync: load REFLEXION.xlsx (or any path) into PostgreSQL.
Used by seed script and OneDrive sync.
Replace mode: truncates tables before load to avoid duplicates.
"""

from __future__ import annotations

import logging
import re
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Optional

import openpyxl
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.estran import EstranRecord
from app.models.finance import FinanceLine
from app.models.purchase import PurchaseDA, PurchaseBC
from app.models.dimensions import DimPeriod, DimEntity

logger = logging.getLogger(__name__)

# ── Header-based column mappings ──────────────────────

PRIMAIRE_HEADER_MAP: dict[str, str] = {
    "génération de semi": "generation_semi",
    "n° ligne": "ligne_num",
    "longueur ligne": "longueur_ligne",
    "orientation w -> e": "orientation",
    "orientation w→e": "orientation",
    "effectif semé (équiv. 200m)": "effectif_seme",
    "effectif semé": "effectif_seme",
    "taille semé": "taille_seme",
    "objectif récolte": "objectif_recolte",
    "objectif recolte": "objectif_recolte",
    "date récolte": "date_recolte",
    "date recolte": "date_recolte",
    "age td (mois)": "age_td_mois",
    "résidence estran (mois)": "residence_estran",
    "residence estran (mois)": "residence_estran",
    "v (kg)": "v_kg",
    "v (kg) /200m": "biomasse_vendable_kg",
    "v (kg)/200m": "biomasse_vendable_kg",
    "tot (kg)": "quantite_brute_recoltee_kg",
    "kg récolté/m²": "kg_recolte_m2",
    "kg recolte/m²": "kg_recolte_m2",
    "kg récolté/m2": "kg_recolte_m2",
    "pm tot (g)": "biomasse_gr",
    "poids mortalité (kg)": "poids_mortalite_kg",
    "poids mortalite (kg)": "poids_mortalite_kg",
    "taux de recapture %": "taux_recapture",
    "taux de recapture": "taux_recapture",
    "nombre de ligne semé (200m)": "nb_ligne_semee_200m",
    "nombre de ligne semee (200m)": "nb_ligne_semee_200m",
    "nb ligne semé 200m": "nb_ligne_semee_200m",
    "nb ligne semee 200m": "nb_ligne_semee_200m",
    "nb ligne semé (200m)": "nb_ligne_semee_200m",
    "nb ligne semee (200m)": "nb_ligne_semee_200m",
    "nb lignes semées 200m": "nb_ligne_semee_200m",
    "nb lignes semees 200m": "nb_ligne_semee_200m",
}

HC_HEADER_MAP: dict[str, str] = {
    "parc de ressemis": "parc_semi",
    "n° ligne": "ligne_num",
    "effectif semé (équiv. 200m)": "effectif_seme",
    "effectif semé": "effectif_seme",
    "effectif seme (équiv. 200m)": "effectif_seme",
    "effectif seme": "effectif_seme",
    "effectif semee (équiv. 200m)": "effectif_seme",
    "effectif semee": "effectif_seme",
    "effectif semé équivalent 200m": "effectif_seme",
    "effectif seme equiv 200m": "effectif_seme",
    "effectif semé equiv 200m": "effectif_seme",
    "v (kg)": "v_kg",
    "origine de récolte prim": "origine",
    "origine de recolte prim": "origine",
    "origine récolte prim": "origine",
    "orientation lignes ouest -> est": "orientation_lignes",
    "orientation lignes ouest→est": "orientation_lignes",
    "orientation o→e": "orientation_lignes",
    "lng de ln semé (m)": "longueur_ligne",
    "lng de ln seme (m)": "longueur_ligne",
    "taille de semi hc": "taille_semi_hc",
    # Biomasse % : somme(Total récolté) / somme(HC Ressemé) — besoin des deux en import HC.
    "total récolté (kg)": "quantite_brute_recoltee_kg",
    "total recolte (kg)": "quantite_brute_recoltee_kg",
    "total récolté kg": "quantite_brute_recoltee_kg",
    "total recolte kg": "quantite_brute_recoltee_kg",
    "tot (kg)": "quantite_brute_recoltee_kg",
    "tot kg": "quantite_brute_recoltee_kg",
    "quantité brute récoltée (kg)": "quantite_brute_recoltee_kg",
    "quantite brute recoltee (kg)": "quantite_brute_recoltee_kg",
    "total récolté": "quantite_brute_recoltee_kg",
    "v (kg) /200m": "biomasse_vendable_kg",
    "v (kg)/200m": "biomasse_vendable_kg",
    "v (kg) / 200m": "biomasse_vendable_kg",
    "v kg/200m": "biomasse_vendable_kg",
    "vendable kg/200m": "biomasse_vendable_kg",
    "biomasse vendable (kg)/200m": "biomasse_vendable_kg",
    "biomasse vendable kg/200m": "biomasse_vendable_kg",
    "hc ressemé (kg)": "quantite_semee_kg",
    "hc resseme (kg)": "quantite_semee_kg",
    "hc ressemé": "quantite_semee_kg",
    "hc resseme": "quantite_semee_kg",
    "hc ressemé kg": "quantite_semee_kg",
    "hc resseme kg": "quantite_semee_kg",
    "hc ressemé: kg/m2": "hc_resseme_kg_m2",
    "hc resseme: kg/m2": "hc_resseme_kg_m2",
    "objectif de récolte": "objectif_recolte",
    "objectif de recolte": "objectif_recolte",
    "date de récolte": "date_recolte",
    "date de recolte": "date_recolte",
    "% biomasse récupérée": "pct_biomasse_recuperee",
    "% biomasse recuperee": "pct_biomasse_recuperee",
    "pm total": "biomasse_gr",
    "mortalité (kg)": "mortalite_kg",
    "mortalite (kg)": "mortalite_kg",
    "% de recapture": "taux_recapture",
    "nombre de ligne semé (200m)": "nb_ligne_semee_200m",
    "nombre de ligne semee (200m)": "nb_ligne_semee_200m",
    "nb ligne semé 200m": "nb_ligne_semee_200m",
    "nb ligne semee 200m": "nb_ligne_semee_200m",
    "nb ligne semé (200m)": "nb_ligne_semee_200m",
    "nb ligne semee (200m)": "nb_ligne_semee_200m",
}

PRIMAIRE_DETECT_KEYS = {"effectif semé", "v (kg) /200m", "pm tot (g)"}
HC_DETECT_KEYS_A = {"hc ressemé (kg)", "pm total", "% de recapture"}
HC_DETECT_KEYS_B = {"hc ressemé (kg)", "pm total", "% biomasse récupérée"}

DATE_FIELDS = {"date_recolte", "date_semis"}
PCT_FIELDS = {"taux_recapture", "pct_biomasse_recuperee"}
INT_FIELDS = {"ligne_num"}
STR_FIELDS = {
    "generation_semi", "orientation", "taille_seme", "objectif_recolte",
    "orientation_lignes", "taille_semi_hc", "parc_semi", "origine",
}


def _norm_header(h: str) -> str:
    return re.sub(r"\s+", " ", str(h).strip()).lower()


def _detect_sheet_type(ws) -> Optional[str]:
    """Return 'primaire', 'hc', or None based on row 1 headers."""
    headers: set[str] = set()
    row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not row1:
        return None
    for cell in row1:
        if cell is not None:
            headers.add(_norm_header(cell))
    if PRIMAIRE_DETECT_KEYS.issubset(headers):
        return "primaire"
    if HC_DETECT_KEYS_A.issubset(headers) or HC_DETECT_KEYS_B.issubset(headers):
        return "hc"
    return None


def _workbook_might_contain_estran(wb) -> bool:
    """True if sheet names or row-1 headers suggest an Estran / REFLEXION workbook."""
    for name in wb.sheetnames:
        ln = str(name).strip().lower()
        if ln in ("primaire", "hors calibre") or "bd estra" in ln:
            return True
    for sname in wb.sheetnames:
        if _detect_sheet_type(wb[sname]) is not None:
            return True
    return False


def _build_header_map(ws, mapping: dict[str, str]) -> dict[int, str]:
    """Map column indices to DB field names based on row 1 headers."""
    col_map: dict[int, str] = {}
    row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not row1:
        return col_map
    for idx, cell in enumerate(row1):
        if cell is None:
            continue
        normed = _norm_header(cell)
        db_field = mapping.get(normed)
        if db_field:
            col_map[idx] = db_field
        else:
            for key, field in mapping.items():
                if key in normed or normed in key:
                    col_map[idx] = field
                    break
    return col_map


def _clean_value(val, field_name: str):
    """Clean a raw Excel cell value based on the target DB field."""
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return None
    if isinstance(val, str) and val.strip().upper() in ("#VALUE!", "#REF!", "#DIV/0!", "NC", "N/A"):
        return None

    if field_name in DATE_FIELDS:
        return _parse_date_flexible(val)

    if field_name in PCT_FIELDS:
        return _parse_percentage(val)

    if field_name in INT_FIELDS:
        return _safe_int(val)

    if field_name in STR_FIELDS:
        s = str(val).strip()
        return s if s else None

    return _safe_decimal(val)


def _parse_date_flexible(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, (int, float)):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=float(val))).date()
        except (ValueError, OverflowError):
            return None
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    logger.warning("Date non parsée: %s", s)
    return None


def _parse_percentage(val) -> Optional[Decimal]:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        v = float(val)
        if 0 < v <= 1.0:
            return Decimal(str(round(v * 100, 4)))
        return Decimal(str(round(v, 4)))
    s = str(val).strip().rstrip("%").strip().replace(",", ".")
    try:
        v = float(s)
        if 0 < v <= 1.0 and "%" not in str(val):
            return Decimal(str(round(v * 100, 4)))
        return Decimal(str(round(v, 4)))
    except (ValueError, TypeError):
        return None


def _derive_year_month(rec: dict) -> None:
    primary = rec.get("date_recolte") or rec.get("date_semis")
    if primary:
        rec["year"] = primary.year
        rec["month"] = primary.month
    else:
        rec["year"] = None
        rec["month"] = None


def _load_sheet_by_headers(ws, col_map: dict[int, str], sheet_label: str) -> list[dict]:
    """Read data rows using header-mapped columns, return list of dicts for EstranRecord."""
    records: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        has_data = False
        rec: dict = {"sheet_name": sheet_label}
        for idx, db_field in col_map.items():
            if idx < len(row):
                val = _clean_value(row[idx], db_field)
                if val is not None:
                    has_data = True
                rec[db_field] = val
        if has_data:
            _derive_year_month(rec)
            records.append(rec)
    return records


# ── Legacy helpers ────────────────────────────────────

def _excel_date_to_python(val):
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    try:
        return (datetime(1899, 12, 30) + timedelta(days=float(val))).date()
    except (TypeError, ValueError):
        return None


def _safe_decimal(val):
    if val is None or val == "" or (isinstance(val, str) and val.upper() in ("#VALUE!", "#REF!", "#DIV/0!", "NC")):
        return None
    try:
        return Decimal(str(val).replace(",", "."))
    except (ValueError, TypeError, InvalidOperation):
        return None


def _safe_int(val):
    if val is None or val == "":
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def _validate_excel_extension(path: Path) -> None:
    ext = path.suffix.lower()
    if ext not in (".xlsx", ".xlsm"):
        raise ValueError(f"Format non autorisé. Utilisez .xlsx ou .xlsm uniquement (reçu: {ext})")


async def seed_from_excel(session: AsyncSession, excel_path: Path, replace: bool = True) -> dict[str, int]:
    """
    Load Excel at excel_path and sync to DB. Returns counts.
    replace=True: truncates tables before load (use for sync).
    replace=False: appends only (use for seed without clearing).
    Only .xlsx and .xlsm are accepted (macro safety).
    """
    _validate_excel_extension(excel_path)
    counts = {"estran": 0, "finance": 0, "purchases": 0}
    counts["primaire_rows"] = 0
    counts["hc_rows"] = 0

    wb_preview = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    try:
        might_estran = _workbook_might_contain_estran(wb_preview)
        has_finance_sheet = "RESULTAT MODELE" in wb_preview.sheetnames
    finally:
        wb_preview.close()

    if replace:
        await session.execute(text("TRUNCATE TABLE purchase_da RESTART IDENTITY CASCADE"))
        await session.execute(text("TRUNCATE TABLE purchase_bc RESTART IDENTITY CASCADE"))
        if might_estran:
            await session.execute(text("TRUNCATE TABLE estran_records RESTART IDENTITY CASCADE"))
        if has_finance_sheet:
            await session.execute(text("TRUNCATE TABLE finance_lines RESTART IDENTITY CASCADE"))
        await session.commit()

    # Dimensions
    r = await session.execute(select(DimPeriod).limit(1))
    if not r.scalar_one_or_none():
        session.add(DimPeriod(year=2025, month=12, label="2025-12"))
        await session.flush()
        session.add(DimEntity(code="SITE1", name="Site Principal", active=True))
        await session.commit()

    # Estran: only when workbook looks like REFLEXION / Primaire+HC (skip for finance-only / achats files)
    if not might_estran:
        logger.debug(
            "Import sans signaux Estran (feuilles Primaire/HC/BD ESTRA ou headers reconnus absents) — "
            "chargement Estran ignoré (normal pour finance/achats)."
        )
    else:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        primaire_loaded = False
        hc_loaded = False
        try:
            for sname in wb.sheetnames:
                ws = wb[sname]
                sheet_type = _detect_sheet_type(ws)

                if sheet_type == "primaire" and not primaire_loaded:
                    col_map = _build_header_map(ws, PRIMAIRE_HEADER_MAP)
                    if col_map:
                        records = _load_sheet_by_headers(ws, col_map, "Primaire")
                        for rec in records:
                            session.add(EstranRecord(**rec))
                        counts["estran"] += len(records)
                        counts["primaire_rows"] = len(records)
                        primaire_loaded = True
                        logger.info("Primaire chargé par headers depuis '%s': %d lignes", sname, len(records))

                elif sheet_type == "hc" and not hc_loaded:
                    col_map = _build_header_map(ws, HC_HEADER_MAP)
                    if col_map:
                        records = _load_sheet_by_headers(ws, col_map, "Hors calibre")
                        for rec in records:
                            session.add(EstranRecord(**rec))
                        counts["estran"] += len(records)
                        counts["hc_rows"] = len(records)
                        hc_loaded = True
                        logger.info("HC chargé par headers depuis '%s': %d lignes", sname, len(records))

                elif sheet_type is None:
                    logger.info("Feuille ignorée (pas de headers reconnus): %s", sname)

            # Legacy fallback: if headers not detected, try position-based loading
            if not primaire_loaded and "Primaire" in wb.sheetnames:
                ws = wb["Primaire"]
                for row in list(ws.iter_rows(min_row=2, values_only=True)):
                    if not row or row[0] is None:
                        continue
                    d_sem = _excel_date_to_python(row[17]) if len(row) > 17 else None
                    d_rec = _excel_date_to_python(row[30]) if len(row) > 30 else None
                    primary = d_rec or d_sem
                    if primary:
                        y_leg, m_leg = primary.year, primary.month
                    else:
                        y_leg = _safe_int(row[97]) if len(row) > 97 else None
                        m_leg = _safe_int(row[98]) if len(row) > 98 else None
                    session.add(EstranRecord(
                        parc_semi=str(row[0]) if row[0] else None,
                        parc_an=str(row[1]) if row[1] else None,
                        generation_semi=str(row[2]) if row[2] else None,
                        ligne_num=_safe_int(row[3]),
                        ett=str(row[4]) if row[4] else None,
                        phase=str(row[5]) if row[5] else None,
                        origine=str(row[6]) if row[6] else None,
                        type_semi=str(row[7]) if row[7] else None,
                        longueur_ligne=_safe_decimal(row[8]) if len(row) > 8 else None,
                        nb_ligne_semee_200m=_safe_decimal(row[9]) if len(row) > 9 else None,
                        zone=str(row[12]) if len(row) > 12 else None,
                        date_semis=d_sem,
                        effectif_seme=_safe_decimal(row[20]) if len(row) > 20 else None,
                        quantite_semee_kg=_safe_decimal(row[23]) if len(row) > 23 else None,
                        quantite_brute_recoltee_kg=_safe_decimal(row[39]) if len(row) > 39 else None,
                        quantite_casse_kg=_safe_decimal(row[40]) if len(row) > 40 else None,
                        biomasse_gr=_safe_decimal(row[63]) if len(row) > 63 else None,
                        biomasse_vendable_kg=_safe_decimal(row[99]) if len(row) > 99 else None,
                        statut=None,
                        etat_recolte=str(row[28]) if len(row) > 28 and row[28] else None,
                        pct_recolte=_safe_decimal(row[31]) if len(row) > 31 else None,
                        date_recolte=d_rec,
                        year=y_leg,
                        month=m_leg,
                        sheet_name="Primaire",
                        type_recolte=str(row[36]).strip() if len(row) > 36 and row[36] else None,
                        taux_recapture=_safe_decimal(row[78]) if len(row) > 78 else None,
                        objectif_recolte=str(row[29]).strip()[:100] if len(row) > 29 and row[29] else None,
                    ))
                    counts["estran"] += 1
                    counts["primaire_rows"] = counts.get("primaire_rows", 0) + 1
                primaire_loaded = True

            if not hc_loaded and "Hors calibre" in wb.sheetnames:
                ws = wb["Hors calibre"]
                for row in list(ws.iter_rows(min_row=2, values_only=True)):
                    if not row or row[0] is None:
                        continue
                    d_sem = _excel_date_to_python(row[0]) if len(row) > 0 else None
                    d_rec = _excel_date_to_python(row[29]) if len(row) > 29 else None
                    primary = d_rec or d_sem
                    if primary:
                        y_hc, m_hc = primary.year, primary.month
                    else:
                        y_hc = _safe_int(row[76]) if len(row) > 76 else None
                        m_hc = _safe_int(row[77]) if len(row) > 77 else None
                    session.add(EstranRecord(
                        parc_semi=str(row[1]) if len(row) > 1 and row[1] else None,
                        parc_an=str(row[2]) if len(row) > 2 and row[2] else None,
                        generation_semi=None, ligne_num=None,
                        ett=str(row[3]) if len(row) > 3 and row[3] else None,
                        phase=str(row[7]) if len(row) > 7 and row[7] else None,
                        origine=str(row[8]) if len(row) > 8 and row[8] else None,
                        type_semi=str(row[6]) if len(row) > 6 and row[6] else None,
                        longueur_ligne=_safe_decimal(row[12]) if len(row) > 12 else None,
                        nb_ligne_semee_200m=_safe_decimal(row[15]) if len(row) > 15 else None,
                        zone=str(row[10]) if len(row) > 10 and row[10] else None,
                        date_semis=d_sem,
                        effectif_seme=_safe_decimal(row[24]) if len(row) > 24 else None,
                        quantite_semee_kg=None,
                        quantite_brute_recoltee_kg=_safe_decimal(row[37]) if len(row) > 37 else None,
                        quantite_casse_kg=_safe_decimal(row[40]) if len(row) > 40 else None,
                        biomasse_gr=None,
                        v_kg=_safe_decimal(row[78]) if len(row) > 78 else None,
                        biomasse_vendable_kg=None,
                        statut=str(row[21]) if len(row) > 21 and row[21] else None,
                        etat_recolte=str(row[27]) if len(row) > 27 and row[27] else None,
                        pct_recolte=_safe_decimal(row[30]) if len(row) > 30 else None,
                        date_recolte=d_rec,
                        year=y_hc,
                        month=m_hc,
                        sheet_name="Hors calibre",
                        type_recolte=str(row[35])[:80] if len(row) > 35 and row[35] else None,
                        taux_recapture=_safe_decimal(row[68]) if len(row) > 68 else None,
                        objectif_recolte=str(row[28])[:100] if len(row) > 28 and row[28] else None,
                    ))
                    counts["estran"] += 1
                    counts["hc_rows"] = counts.get("hc_rows", 0) + 1
                hc_loaded = True

            elif not hc_loaded and not primaire_loaded and "BD ESTRA" in wb.sheetnames:
                ws = wb["BD ESTRA"]
                for row in list(ws.iter_rows(min_row=2, values_only=True)):
                    if not row or row[0] is None:
                        continue
                    d_sem = _excel_date_to_python(row[17]) if len(row) > 17 else None
                    d_rec = _excel_date_to_python(row[29]) if len(row) > 29 else None
                    primary = d_rec or d_sem
                    if primary:
                        y_bd, m_bd = primary.year, primary.month
                    else:
                        y_bd = _safe_int(row[81]) if len(row) > 81 else None
                        m_bd = _safe_int(row[82]) if len(row) > 82 else None
                    session.add(EstranRecord(
                        parc_semi=str(row[0]) if row[0] else None,
                        parc_an=str(row[1]) if row[1] else None,
                        generation_semi=str(row[2]) if row[2] else None,
                        ligne_num=_safe_int(row[3]),
                        ett=str(row[4]) if row[4] else None,
                        phase=str(row[5]) if row[5] else None,
                        origine=str(row[6]) if row[6] else None,
                        type_semi=str(row[7]) if row[7] else None,
                        longueur_ligne=_safe_decimal(row[8]) if len(row) > 8 else None,
                        nb_ligne_semee_200m=_safe_decimal(row[9]) if len(row) > 9 else None,
                        zone=str(row[12]) if len(row) > 12 else None,
                        date_semis=d_sem,
                        effectif_seme=_safe_decimal(row[20]) if len(row) > 20 else None,
                        quantite_semee_kg=_safe_decimal(row[23]) if len(row) > 23 else None,
                        quantite_brute_recoltee_kg=_safe_decimal(row[36]) if len(row) > 36 else None,
                        quantite_casse_kg=_safe_decimal(row[37]) if len(row) > 37 else None,
                        biomasse_gr=_safe_decimal(row[79]) if len(row) > 79 else None,
                        biomasse_vendable_kg=_safe_decimal(row[83]) if len(row) > 83 else None,
                        statut=str(row[76]) if len(row) > 76 else None,
                        etat_recolte=str(row[28]) if len(row) > 28 else None,
                        pct_recolte=_safe_decimal(row[30]) if len(row) > 30 else None,
                        date_recolte=d_rec,
                        year=y_bd,
                        month=m_bd,
                        sheet_name=None, type_recolte=None, taux_recapture=None, objectif_recolte=None,
                    ))
                    counts["estran"] += 1

            if not primaire_loaded:
                logger.warning("Aucune feuille Primaire détectée dans le fichier")
            if not hc_loaded:
                logger.warning("Aucune feuille HC détectée dans le fichier")
        finally:
            wb.close()

    await session.commit()

    # Finance
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    if "RESULTAT MODELE" in wb.sheetnames:
            ws = wb["RESULTAT MODELE"]
            for row in list(ws.iter_rows(min_row=2, values_only=True))[:100]:
                if not row or row[0] is None:
                    continue
                session.add(FinanceLine(
                    code=str(row[0]).strip() if row[0] else "",
                    ordre=_safe_int(row[1]),
                    gr=str(row[2]).strip() if row[2] else None,
                    label=(str(row[3])[:255] if row[3] is not None else None),
                    n1=_safe_decimal(row[4]),
                    budget=_safe_decimal(row[5]),
                    real=_safe_decimal(row[7]),
                    fy=_safe_decimal(row[8]),
                    var_b_r=_safe_decimal(row[9]),
                    var_pct=_safe_decimal(row[10]),
                    var_r_n1=_safe_decimal(row[11]) if len(row) > 11 else None,
                    ytd=_safe_decimal(row[6]) if len(row) > 6 and isinstance(row[6], (int, float)) else None,
                    year=2025,
                    month=12,
                ))
                counts["finance"] += 1
    wb.close()

    # Purchases : lignes de démo uniquement si le classeur a bien alimenté Estran ou Finance
    # (évite d'afficher « que des achats » après import d'un mauvais fichier ex. Suivi CCS seul).
    if counts["estran"] > 0 or counts["finance"] > 0:
        da_bc = [
            *[PurchaseDA(reference="DA-2025-001", amount=Decimal("15000"), delay_days=5, status="En cours", critical_flag=False),
              PurchaseDA(reference="DA-2025-002", amount=Decimal("45000"), delay_days=12, status="En cours", critical_flag=True)],
            *[PurchaseBC(reference="BC-2025-101", amount=Decimal("22000"), delay_days=3, status="Non livré", critical_flag=False, expected_delivery_date=date(2025, 2, 15)),
              PurchaseBC(reference="BC-2025-102", amount=Decimal("78000"), delay_days=18, status="Non livré", critical_flag=True, expected_delivery_date=date(2025, 1, 28))],
        ]
        for r in da_bc:
            session.add(r)
            counts["purchases"] += 1
    await session.commit()

    return counts
